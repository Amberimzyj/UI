# -*- coding: utf-8 -*-
# @Author: Amberimzyj
# @Emial:  amberimzyj@outlook.com
# @Date:   2018-05-01 15:36:48
# @Last Modified by:   Amberimzyj
# @Last Modified time: 2018-05-03 20:19:45
# @License: MIT LICENSE
from openpyxl import load_workbook,Workbook



class Xlsx(object):
    def __init__(self):
        # self._sheet1_count = 1
        # self._sheet2_count = 1
        # self._sheet2_movie_count = 0
        # 初始化工作簿
        # self._xlsx_wb = load_workbook('D:\study\The eighth term\graduation design\data\Measured_data.xlsx')
        self.wb = Workbook()
        self._xlsx_ws1 = self.wb.active
        self._xlsx_ws1.title = "F1"
        self._xlsx_ws2 = self.wb.create_sheet(title="F2")
        self._xlsx_ws3 = self.wb.create_sheet(title="F3")
        self._xlsx_ws4 = self.wb.create_sheet(title="F4")

        # self._xlsx_ws1 = self._xlsx_wb.get_sheet_by_name(sheetnames[0])
        # self._xlsx_ws2 = self._xlsx_wb.get_sheet_by_name(sheetnames[1])
        # self._xlsx_ws3 = self._xlsx_wb.get_sheet_by_name(sheetnames[2])
        # self._xlsx_ws4 = self._xlsx_wb.get_sheet_by_name(sheetnames[3])


    def __del__(self):
        pass
        # del self._sheet1_count
        # del self._sheet2_count
        # del self._sheet2_movie_count

    def set_sheet1_data(self,_data):
        # self._sheet1_count += 1
        self._xlsx_ws1.cell(row=1,column=1,value='时间')
        self._xlsx_ws1.cell(row=1,column=2,value='温度(℃)')
        self._xlsx_ws1.cell(row=1,column=3,value='湿度(%RH)')
        self._xlsx_ws1.cell(row=1,column=4,value='光照强度(lx)')
        # len1=self.l1+3

        for len1 in range(len(_data)):
            self._xlsx_ws1.cell(row=len1+2,column=1,value=_data[len1][0])
            self._xlsx_ws1.cell(row=len1+2,column=2,value=_data[len1][1])
            self._xlsx_ws1.cell(row=len1+2,column=3,value=_data[len1][2])
            self._xlsx_ws1.cell(row=len1+2,column=4,value=_data[len1][3])
            # self.l1 += 4
            # len1=self.l1+3

    def set_sheet2_data(self,_data):
        # self._sheet1_count += 1
        self._xlsx_ws2.cell(row=1,column=1,value='时间')
        self._xlsx_ws2.cell(row=1,column=2,value='温度(℃)')
        self._xlsx_ws2.cell(row=1,column=3,value='湿度(%RH)')
        self._xlsx_ws2.cell(row=1,column=4,value='光照强度(lx)')
        # len1=self.l1+3

        for len1 in range(len(_data)):
            self._xlsx_ws2.cell(row=len1+2,column=1,value=_data[len1][0])
            self._xlsx_ws2.cell(row=len1+2,column=2,value=_data[len1][1])
            self._xlsx_ws2.cell(row=len1+2,column=3,value=_data[len1][2])
            self._xlsx_ws2.cell(row=len1+2,column=4,value=_data[len1][3])

    def set_sheet3_data(self,_data):
        # self._sheet1_count += 1
        self._xlsx_ws3.cell(row=1,column=1,value='时间')
        self._xlsx_ws3.cell(row=1,column=2,value='温度(℃)')
        self._xlsx_ws3.cell(row=1,column=3,value='湿度(%RH)')
        self._xlsx_ws3.cell(row=1,column=4,value='光照强度(lx)')
        # len1=self.l1+3

        for len1 in range(len(_data)):
            self._xlsx_ws3.cell(row=len1+2,column=1,value=_data[len1][0])
            self._xlsx_ws3.cell(row=len1+2,column=2,value=_data[len1][1])
            self._xlsx_ws3.cell(row=len1+2,column=3,value=_data[len1][2])
            self._xlsx_ws3.cell(row=len1+2,column=4,value=_data[len1][3])

    def set_sheet4_data(self,_data):
        # self._sheet1_count += 1
        self._xlsx_ws4.cell(row=1,column=1,value='时间')
        self._xlsx_ws4.cell(row=1,column=2,value='温度(℃)')
        self._xlsx_ws4.cell(row=1,column=3,value='湿度(%RH)')
        self._xlsx_ws4.cell(row=1,column=4,value='光照强度(lx)')
        # len1=self.l1+3

        for len1 in range(len(_data)):
            self._xlsx_ws4.cell(row=len1+2,column=1,value=_data[len1][0])
            self._xlsx_ws4.cell(row=len1+2,column=2,value=_data[len1][1])
            self._xlsx_ws4.cell(row=len1+2,column=3,value=_data[len1][2])
            self._xlsx_ws4.cell(row=len1+2,column=4,value=_data[len1][3])

    # def save_workbook(self):
    #     self._xlsx_wb.save('Measured_data.xlsx')


if __name__ == '__main__':
    test = Xlsx()
    # test.save_workbook()
